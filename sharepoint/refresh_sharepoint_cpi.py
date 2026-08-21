
import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_OUTPUT = Path(__file__).parent / "output" / "IMF - Consumer Price Index (CPI).xlsx"

EXCEL_SAFETY_MARGIN = 900_000

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

METADATA_COLUMN_MAP = {
    "dataset_name": "Dataset name",
    "dataset_id": "ID",
    "agency": "Agency",
    "version": "Version",
    "dataset_description": "Dataset Description",
    "geographical_coverage": "Geographical Coverage",
    "full_description": "Full Description",
    "publisher": "Publisher",
    "department": "Department",
    "contact_point": "Contact Point",
    "topic_dataset": "Topic Dataset",
    "keywords_dataset": "Keywords Dataset",
    "language": "Language",
    "publication_date": "Publication Date",
    "update_date": "Update Date",
    "short_source_citation": "Short Source Citation",
    "full_source_citation": "Full Source Citation",
    "license": "License",
    "suggested_citation": "Suggested Citation",
}

DATA_TRIPLET_DIMENSIONS = [
    "country", "index_type", "coicop_1999", "type_of_transformation",
    "frequency", "scale", "precision", "decimals_displayed",
    "reporting_period_type", "transformation", "unit", "derivation_type",
    "overlap", "publisher", "department", "topic", "topic_dataset",
    "language", "methodology", "access_sharing_level",
    "security_classification", "source",
]

DATA_STANDALONE_COLUMNS = [
    "time_period", "obs_value", "reference_period", "common_reference_period",
    "status", "ifs_flag", "doi", "full_description", "author",
    "contact_point", "keywords", "keywords_dataset", "publication_date",
    "update_date", "methodology_notes", "access_sharing_notes",
    "short_source_citation", "full_source_citation", "license",
    "suggested_citation", "key_indicator", "series_name",
]


def build_data_column_map():
    mapping = {}
    for base in DATA_TRIPLET_DIMENSIONS:
        upper = base.upper()
        mapping[f"{base}_id"] = f"{upper}.ID"
        mapping[base] = upper
        mapping[f"{base}_description"] = f"{upper}.Description"
    for col in DATA_STANDALONE_COLUMNS:
        mapping[col] = col.upper()
    return mapping


DATA_COLUMN_MAP = build_data_column_map()


def fetch_metadata(dsn: str):
    import psycopg2
    import psycopg2.extras

    columns = ", ".join(METADATA_COLUMN_MAP.keys())
    conn = psycopg2.connect(dsn)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(f"SELECT {columns} FROM cpi.metadata ORDER BY dataset_id;")
            return cur.fetchall()
    finally:
        conn.close()


def get_latest_successful_run(dsn: str, request_mode: str = "API"):
    """Retourne le dernier run réussi. request_mode='API' correspond à
    collect_cpi_data.py (voir CHECK chk_logs_request_mode : 'API' ou
    'Web Scraping' — pas de colonne 'pipeline' ni de valeur
    'collect_cpi_data' dans le schéma corrigé)."""
    import psycopg2

    conn = psycopg2.connect(dsn)
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, end_date, persisted_elements_count
                FROM cpi.logs
                WHERE request_mode = %s AND status = 'SUCCESS'
                ORDER BY end_date DESC
                LIMIT 1;
                """,
                (request_mode,),
            )
            return cur.fetchone()
    finally:
        conn.close()


def fetch_data(dsn: str, logs_id: int, coicop_filter=None):
    import psycopg2
    import psycopg2.extras

    columns = ", ".join(DATA_COLUMN_MAP.keys())
    conn = psycopg2.connect(dsn)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            query = f"SELECT {columns} FROM cpi.data WHERE logs_id = %s"
            params = [logs_id]
            if coicop_filter:
                query += " AND coicop_1999_id = ANY(%s)"
                params.append(coicop_filter)
            query += " ORDER BY country, coicop_1999, time_period;"
            cur.execute(query, params)
            return cur.fetchall()
    finally:
        conn.close()


def rename_rows(rows, column_map):
    renamed = []
    for row in rows:
        renamed.append({column_map[k]: v for k, v in row.items()})
    return renamed


def write_sheet(ws, rows, headers=None):
    if not rows:
        ws.append(["Aucune donnée disponible"])
        return

    if headers is None:
        headers = list(rows[0].keys())

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row in rows:
        ws.append([row.get(h) for h in headers])

    for col_idx, header in enumerate(headers, start=1):
        max_len = max(
            [len(str(header))] + [len(str(row.get(header, ""))) for row in rows if row.get(header) is not None]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"


def write_fields_sheet(ws, data_headers):
    ws.append(["Fields"])
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT
    for header in data_headers:
        ws.append([header])
    ws.column_dimensions["A"].width = 40


def build_workbook(metadata_rows, data_rows):
    wb = Workbook()

    metadata_headers = list(METADATA_COLUMN_MAP.values())
    metadata_renamed = rename_rows(metadata_rows, METADATA_COLUMN_MAP)
    ws_meta = wb.active
    ws_meta.title = "Metadata"
    write_sheet(ws_meta, metadata_renamed, headers=metadata_headers)

    data_headers = list(DATA_COLUMN_MAP.values())
    data_renamed = rename_rows(data_rows, DATA_COLUMN_MAP)
    ws_data = wb.create_sheet("Data")
    write_sheet(ws_data, data_renamed, headers=data_headers)

    ws_fields = wb.create_sheet("Fields")
    write_fields_sheet(ws_fields, data_headers)
    wb.move_sheet("Fields", offset=-1)

    return wb


def refresh_sharepoint_file(dsn: str, output_path: Path, coicop_filter):
    print("Connexion à la base et lecture de cpi.metadata...")
    metadata_rows = fetch_metadata(dsn)
    print(f"  {len(metadata_rows)} ligne(s) metadata.")

    print("Recherche du dernier run réussi (request_mode='API')...")
    latest_run = get_latest_successful_run(dsn)
    if not latest_run:
        print("Erreur : aucun run réussi trouvé dans cpi.logs (request_mode='API').", file=sys.stderr)
        sys.exit(1)
    logs_id, end_date, nb_persistes = latest_run
    print(f"  Dernier run : logs_id={logs_id} (terminé le {end_date}, {nb_persistes} ligne(s))")

    if coicop_filter:
        print(f"⚠️  Filtre temporaire actif : coicop_1999 IN {coicop_filter} "
              f"(voir README — périmètre à valider avec l'équipe)")

    print("Lecture de cpi.data pour ce run...")
    data_rows = fetch_data(dsn, logs_id, coicop_filter)
    print(f"  {len(data_rows)} ligne(s) data.")

    if len(data_rows) > EXCEL_SAFETY_MARGIN:
        print(
            f"❌ {len(data_rows)} lignes dépassent la marge de sécurité Excel "
            f"({EXCEL_SAFETY_MARGIN}). Fichier NON généré — resserrer le filtre "
            f"(coicop_filter) ou attendre la décision d'équipe sur le périmètre.",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = build_workbook(metadata_rows, data_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Fichier régénéré : {output_path}")

    return len(metadata_rows), len(data_rows)


def main():
    parser = argparse.ArgumentParser(
        description="Alimente le fichier SharePoint CPI depuis la base PostgreSQL (dernier run uniquement)."
    )
    parser.add_argument("--dsn", required=True, help="Chaîne de connexion PostgreSQL")
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help="Chemin de destination (dossier OneDrive/SharePoint synchronisé localement).",
    )
    parser.add_argument(
        "--coicop", nargs="*", default=["_T"],
        help="Filtre temporaire sur les catégories COICOP (défaut : _T uniquement).",
    )
    args = parser.parse_args()

    output_path = Path(args.output)
    coicop_filter = args.coicop if args.coicop else None

    start = datetime.now(timezone.utc)
    try:
        nb_metadata, nb_data = refresh_sharepoint_file(args.dsn, output_path, coicop_filter)
    except PermissionError as e:
        print(
            f"Erreur : impossible d'écrire le fichier. Assure-toi qu'il n'est pas "
            f"ouvert dans Excel. Détail : {e}",
            file=sys.stderr,
        )
        sys.exit(1)
    except SystemExit:
        raise
    except Exception as e:
        print(f"Erreur : {e}", file=sys.stderr)
        sys.exit(1)

    duration = (datetime.now(timezone.utc) - start).total_seconds()
    print(
        f"Résumé : metadata={nb_metadata} ligne(s) | data={nb_data} ligne(s) "
        f"| durée={duration:.1f}s | statut=SUCCESS"
    )


if __name__ == "__main__":
    main()